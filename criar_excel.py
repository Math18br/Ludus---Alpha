from openpyxl import Workbook, load_workbook
import os


def adicionar_dados_excel(nome_arquivo, data_dicionario):
    pasta_arquivos = os.path.join(os.getcwd(), 'arquivos salvos')

    if not os.path.exists(pasta_arquivos):
        os.makedirs(pasta_arquivos)

    caminho_arquivo = os.path.join(pasta_arquivos, nome_arquivo)

    if os.path.exists(caminho_arquivo):
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(data_dicionario.keys()))

    ws.append(list(data_dicionario.values()))
    wb.save(caminho_arquivo)
